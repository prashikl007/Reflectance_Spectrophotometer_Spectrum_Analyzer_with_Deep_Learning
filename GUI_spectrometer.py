import cv2
from tkinter import *
from PIL import Image, ImageTk
import datetime
import openpyxl
import time
import numpy as np
import openpyxl
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import keras

#initialization and loading required data
model_name = "spectrum_model.h5"
model = keras.models.load_model(model_name)
predict_pressed = 0

#here starts required functions
def load_xl_data():
   # loading wavelength data from excel sheet
   start = time.time()
   xl_path = "pixel_wavelength.xlsx"
   wb_obj = openpyxl.load_workbook(xl_path)

   sheet_obj = wb_obj.active

   row = sheet_obj.max_row
   column = sheet_obj.max_column
   print("Total Rows:", row)
   print("Total Columns:", column)

   color_pixel = []
   wavelength = []
   for i in range(0, 315, 1):
      # print('current i = ', i)
      obj1 = sheet_obj.cell(row=1 + i, column=1)
      r = int(obj1.value)
      obj2 = sheet_obj.cell(row=1 + i, column=2)
      g = int(obj2.value)
      obj3 = sheet_obj.cell(row=1 + i, column=3)
      b = int(obj3.value)
      color_pixel.append([r, g, b])
      obj4 = sheet_obj.cell(row=1 + i, column=4)
      wavelength.append(float(obj4.value))

   stop = time.time()
   print("time required to load data from xls = ", stop - start)
   return color_pixel, wavelength


def detect_wavelength(selected_color, selected_color_count, color_pixel, wavelength):
   # detecting wavelength for selected color
   start = time.time()
   detected_wavelenth = []
   wavelength_intensity = []
   # finding rms difference between selected color and standard color
   for i in range(len(selected_color)):
      rms = []
      for j in range(len(color_pixel)):
         # r - red color values
         r_dp = color_pixel[j][0] - selected_color[i][0]
         r_sr = r_dp * r_dp
         # g - green color value
         g_da = color_pixel[j][1] - selected_color[i][1]
         g_ss = g_da * g_da
         # b - blue color value
         b_dh = color_pixel[j][2] - selected_color[i][2]
         b_si = b_dh * b_dh
         mean_k = (r_sr + g_ss + b_si) / 3
         rms.append(np.sqrt(mean_k))

      min_rms = min(rms)
      for k in range(len(rms)):
         if (rms[k] == min_rms):
            detected_wavelenth.append(wavelength[k])
            wavelength_intensity.append(selected_color_count[i])
   stop = time.time()
   print("time required to detect wavelength = ", stop - start)
   return detected_wavelenth, wavelength_intensity
# end of detect wavelength function


def get_non_repeated_wavelength(wavelength, intensity):
   start = time.time()
   final_wavelength = []
   final_wavelength.append(wavelength[0])
   final_intensity = []
   final_intensity.append(intensity[0])
   la = len(wavelength)
   for i in range(1, la, 1):
      l = len(final_wavelength)
      for j in range(l):
         if (final_wavelength[j] == wavelength[i]):
            x = final_intensity[j] + intensity[i]
            final_intensity[j] = x
            result = 1
            break
         else:
            result = 0

      if (result == 0):
         final_wavelength.append(wavelength[i])
         final_intensity.append(intensity[i])
         result = 1

   stop = time.time()
   print("time required to get non repeated wavelength = ", stop - start)
   return final_wavelength, final_intensity
# end of get_non_repeated_wavelength()

intensity = []
def get_plot_from_wavelength(color_pixel, wavelength, win, prediction):
   global predict_pressed
   if predict_pressed == 1:
      prediction.config(text =  "__________")
      
   img_path = 'C:/Users/Public/PYTHON_PROJECTS/spectrometer/Al_foil.jpg'
   #image = cv2.imread(img_path)
   #cv2.imshow("image", image)
   image = cap.read()[1]
   cropped_image = image[80:300,450:610]
   cv2.imshow("image", cropped_image)
   cv2.waitKey(0)
   cv2.destroyAllWindows()
   image = cropped_image
   h, w, _ = image.shape
   print("height = ", h, ", width = ", w)

   total_start = time.time()

   # ordering BGR to RGB from original image into img_rgb
   img_rgb = []
   for i in range(h):
      img_rgb.append([])
      for j in range(w):
         img_rgb[i].append([image[i][j][2], image[i][j][1], image[i][j][0]])
   print("completed ordering")

   unique_color_values = []
   unique_color_values.append(img_rgb[0][0])
   # searching for unique colors and stored in unique_color_values
   start = time.time()
   for i in range(h):
      for j in range(w):
         total_color = len(unique_color_values)
         for k in range(total_color):
            if ((img_rgb[i][j][0] == unique_color_values[k][0]) and (
                    img_rgb[i][j][1] == unique_color_values[k][1]) and (
                    img_rgb[i][j][2] == unique_color_values[k][2])):
               result = 1
               break
            else:
               result = 0

         if result == 0:
            unique_color_values.append(img_rgb[i][j])
            result = 1

   # selecting those color pixel which are greater than 70
   # and remving black and gray color pixels
   unique_color = []
   l = len(unique_color_values)
   for i in range(l):
      if ((unique_color_values[i][0] > 20) or
              (unique_color_values[i][1] > 20) or
              (unique_color_values[i][2] > 20)):
         if ((unique_color_values[i][0] != unique_color_values[i][1]) and
                 (unique_color_values[i][0] != unique_color_values[i][2])):
            unique_color.append(unique_color_values[i])
   stop = time.time()
   print("time required for getting unique colors = ", stop - start)

   # calculating count of unique color values
   color_count = []
   start = time.time()
   for i in range(len(unique_color)):
      count = 0
      for j in range(h):
         for k in range(w):
            if ((unique_color[i][0] == img_rgb[j][k][0]) and (
                    unique_color[i][1] == img_rgb[j][k][1]) and (
                    unique_color[i][2] == img_rgb[j][k][2])):
               count = count + 1
      color_count.append(count)
   stop = time.time()
   print("time required for calculating count of every unique color = ", stop - start)

   # selecting colors which will be significant
   # by removing those colors which have very less count
   avg = np.mean(color_count)
   start = time.time()
   l = len(color_count)
   selected_color = []
   selected_color_count = []
   threshold = 0
   for i in range(l):
      if (color_count[i] > threshold):
         selected_color.append(unique_color[i])
         selected_color_count.append(color_count[i])
   stop = time.time()
   print("time required for selecting colors = ", stop - start)

   detected_wavelength, detected_wavelength_intensity = detect_wavelength(selected_color, selected_color_count,
                                                                          color_pixel, wavelength)
   final_wavelength, final_intensity = get_non_repeated_wavelength(detected_wavelength, detected_wavelength_intensity)

   start = time.time()
   sorted_final_wavelength = sorted(final_wavelength, reverse=True)
   sorted_final_intensity = []
   for i in range(len(sorted_final_wavelength)):
      for j in range(len(final_wavelength)):
         if (sorted_final_wavelength[i] == final_wavelength[j]):
            sorted_final_intensity.append(final_intensity[j])

   global intensity
   intensity = []
   for i in range(len(wavelength)):
      for j in range(len(sorted_final_wavelength)):
         if (wavelength[i] == sorted_final_wavelength[j]):
            intensity.append(sorted_final_intensity[j])
            result = 1
            break
         else:
            result = 0

      if (result == 0):
         intensity.append(0)
         result = 1
   stop = time.time()
   print("time required to sort wavelength and intensity = ", stop - start)

   figure = plt.figure(figsize=(5, 4), dpi=100)
   figure.add_subplot(111).plot(wavelength, intensity)
   chart = FigureCanvasTkAgg(figure, master = win)
   chart.get_tk_widget().grid(row=1, column=1)

   total_stop = time.time()
   B3 = Button(win,text="Store in Dataset",font=("Times new roman",10,"bold"),bg="white",fg="red",command=pop_up).grid(row=1, column=2, )
   print("total time elapsed = ", total_stop - total_start)
# end of function get_plot_from_wavelength()

def predict_material(prediction):
   materials = ["caco3", "nacl", "table_sugar"]
   global intensity
   global model
   x = np.expand_dims(intensity, axis=0)
   new_pred = model.predict(x)
   max_value = max(new_pred[0])
   for i in range(len(materials)):
       if new_pred[0][i] == max_value:
           index = i
   print("detected material = ", materials[index])
   prediction.config(text = materials[index])
   global predict_pressed
   predict_pressed = 1
   


#start of GUI functions
def show_frames():
   im = cap.read()[1]
   im = cv2.rectangle(im, (450, 80), (600, 300), (0, 255, 255), 2)
   Image1= cv2.cvtColor(im,cv2.COLOR_BGR2RGB)
   img = Image.fromarray(Image1)
   imgtk = ImageTk.PhotoImage(image = img)
   label.imgtk = imgtk
   label.configure(image=imgtk)
   label.after(20, show_frames)

def capture():
   I = cap.read()[1]
   save_name = str(datetime.datetime.now().today()).replace(":", " ") + ".jpg"
   cv2.imwrite(save_name, I)

def quit_me():
   print('Quit')
   win.quit()
   win.destroy()


top = 0
def pop_up():
   global top
   top = Toplevel(win)
   top.geometry("150x150")
   element = Label(top,text = "Provide name of element ",font=("times new roman",10,"bold"),bg="white",fg="red")
   element.pack()
   entry = Entry(top, width = 25)
   entry.pack()
   submit =  Button(top, text="Submit", command=lambda:close_pop_up(entry))
   submit.pack()

def close_pop_up(entry):
   xl_path = "spectrum_wavelength_dataset.xlsx"
   wb_obj = openpyxl.load_workbook(xl_path)
   sheet_obj = wb_obj.active
   row = sheet_obj.max_row
   column = sheet_obj.max_column
   next_row = row+1
   start = time.time()
   cell_obj = sheet_obj.cell(row = next_row, column = 1)
   cell_obj.value = entry.get()
   wb_obj.save(xl_path)
   global intensity
   len_intensity = len(intensity)
   for i in range(len_intensity):
       cell_obj = sheet_obj.cell(row = next_row, column = i+2)
       cell_obj.value = intensity[i]
       wb_obj.save(xl_path)
   stop = time.time()
   print("time required to store dataset = ", stop-start)
   global top
   top.destroy()

#here creating window and GUI
win = Tk()
win.protocol("WM_DELETE_WINDOW", quit_me)
win.geometry("1200x650")
win.title('Portable Optical Spectrometer')

#scroll_bar1 = Scrollbar(win)
#scroll_bar1.pack(side = RIGHT, orient = vertical)

color_pixel, wavelength = load_xl_data()



L2 = Label(win,text = " Camera  ",font=("times new roman",10,"bold"),bg="white",fg="red").grid(row=0, column=0)
L3 = Label(win,text = " Spectrum Graph ",font=("times new roman",10,"bold"),bg="white",fg="red").grid(row=0, column=1)
label =Label(win)
label.grid(row=1, column=0)
cap = cv2.VideoCapture(1)

show_frames()
B1 = Button(win,text="Capture",font=("Times new roman",10,"bold"),bg="white",fg="red",command=capture).grid(row=3, column=0, )
B2 = Button(win,text="Analysis",font=("Times new roman",10,"bold"),bg="white",fg="red",command =lambda: get_plot_from_wavelength(color_pixel, wavelength, win, prediction)).grid(row=4, column=0)
prediction = Label(win, text = "__________", height = 3,width = 25,bg = "light cyan")
prediction.grid(row=4, column=2)
B3 = Button(win,text="Predict material",font=("Times new roman",10,"bold"),bg="white",fg="red",command= lambda: predict_material(prediction)).grid(row=3, column=1, )
L1 = Label(win,text = "Detected Material is: ------>  ",font=("times new roman",10,"bold"),bg="white",fg="red").grid(row=4, column=1)


win.mainloop()
cap.release()
