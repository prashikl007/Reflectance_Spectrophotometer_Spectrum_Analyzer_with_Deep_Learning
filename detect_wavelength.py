import cv2
import time
import numpy as np
import openpyxl
import matplotlib.pyplot as plt



def load_xl_data ():
    # loading wavelength data from excel sheet
    start = time.time()
    xl_path = "pixel_wavelength.xlsx"
    wb_obj = openpyxl.load_workbook(xl_path)

    sheet_obj = wb_obj.active

    row = sheet_obj.max_row
    column = sheet_obj.max_column
    print("Total Rows:", row)
    print("Total Columns:", column)

    color_pixel = []
    wavelength = []
    for i in range(row):
        obj1 = sheet_obj.cell(row=1 + i, column=1)
        r = int(obj1.value)
        obj2 = sheet_obj.cell(row=1 + i, column=2)
        g = int(obj2.value)
        obj3 = sheet_obj.cell(row=1 + i, column=3)
        b = int(obj3.value)
        color_pixel.append([r, g, b])
        obj4 = sheet_obj.cell(row=1 + i, column=4)
        wavelength.append(float(obj4.value))

    stop = time.time()
    print("time required to load data from xls = ", stop - start)
    return color_pixel, wavelength


def detect_wavelength(selected_color, selected_color_count, color_pixel, wavelength):
    # detecting wavelength for selected color
    start = time.time()
    detected_wavelenth = []
    wavelength_intensity = []
    # finding rms difference between selected color and standard color
    for i in range(len(selected_color)):
        rms = []
        for j in range(len(color_pixel)):
            #r - red color values
            r_dp = color_pixel[j][0] - selected_color[i][0]
            r_sr = r_dp * r_dp
            #g - green color value
            g_da = color_pixel[j][1] - selected_color[i][1]
            g_ss = g_da * g_da
            #b - blue color value
            b_dh = color_pixel[j][2] - selected_color[i][2]
            b_si = b_dh * b_dh
            mean_k = (r_sr + g_ss + b_si) / 3
            rms.append(np.sqrt(mean_k))

        min_rms = min(rms)
        for k in range(len(rms)):
            if (rms[k] == min_rms):
                detected_wavelenth.append(wavelength[k])
                wavelength_intensity.append(selected_color_count[i])
    stop = time.time()
    print("time required to detect wavelength = ", stop - start)
    return detected_wavelenth, wavelength_intensity


def get_non_repeated_wavelength(wavelength, intensity):
    start = time.time()
    final_wavelength = []
    final_wavelength.append(wavelength[0])
    final_intensity = []
    final_intensity.append(intensity[0])
    la = len(wavelength)
    for i in range(1, la, 1):
        l = len(final_wavelength)
        for j in range(l):
            if (final_wavelength[j] == wavelength[i]):
                x = final_intensity[j] + intensity[i]
                final_intensity[j] = x
                result = 1
                break
            else:
                result = 0

        if (result == 0):
            final_wavelength.append(wavelength[i])
            final_intensity.append(intensity[i])
            result = 1

    stop = time.time()
    print("time required to get non repeated wavelength = ", stop - start)
    return final_wavelength, final_intensity



img_path = 'C:/Users/Public/PYTHON_PROJECTS/spectrometer/spectrum_3.jpg'
image = cv2.imread(img_path)
cv2.imshow("image", image)
cv2.waitKey(0)
cv2.destroyAllWindows()

h, w, _ = image.shape
print("height = ", h , ", width = ", w)

total_start = time.time()

#ordering BGR to RGB from original image into img_rgb
img_rgb = []
for i in range(h):
    img_rgb.append([])
    for j in range(w):
        img_rgb[i].append([image[i][j][2], image[i][j][1], image[i][j][0]])
print("completed ordering")


unique_color_values = []
unique_color_values.append(img_rgb[0][0])
# searching for unique colors and stored in unique_color_values
start = time.time()
for i in range(h):
    for j in range(w):
        total_color = len(unique_color_values)
        for k in range(total_color):
            if ((img_rgb[i][j][0] == unique_color_values[k][0]) and (
                    img_rgb[i][j][1] == unique_color_values[k][1]) and (
                    img_rgb[i][j][2] == unique_color_values[k][2])):
                result = 1
                break
            else:
                result = 0

        if result == 0:
            unique_color_values.append(img_rgb[i][j])
            result = 1

#selecting those color pixel which are greater than 70
#and remving black and gray color pixels
unique_color = []
l=len(unique_color_values)
for i in range(l):
    if ((unique_color_values[i][0] > 70) or
            (unique_color_values[i][1] > 70) or
            (unique_color_values[i][2] > 70)):
        if ((unique_color_values[i][0] != unique_color_values[i][1]) and
                (unique_color_values[i][0] != unique_color_values[i][2])):
            unique_color.append(unique_color_values[i])
stop = time.time()
print("time required for getting unique colors = ", stop - start)

# calculating count of unique color values
color_count = []
start = time.time()
for i in range(len(unique_color)):
    count = 0
    for j in range(h):
        for k in range(w):
            if ((unique_color[i][0] == img_rgb[j][k][0]) and (
                    unique_color[i][1] == img_rgb[j][k][1]) and (
                    unique_color[i][2] == img_rgb[j][k][2])):
                count = count + 1
    color_count.append(count)
stop = time.time()
print("time required for calculating count of every unique color = ", stop - start)

# selecting colors which will be significant
# by removing those colors which have very less count
avg = np.mean(color_count)
start = time.time()
l = len(color_count)
selected_color = []
selected_color_count = []
threshold = 0
for i in range(l):
    if (color_count[i] > threshold):
        selected_color.append(unique_color[i])
        selected_color_count.append(color_count[i])
stop = time.time()
print("time required for selecting colors = ", stop - start)


color_pixel, wavelength =  load_xl_data()

detected_wavelength, detected_wavelength_intensity = detect_wavelength(selected_color, selected_color_count, color_pixel, wavelength)

final_wavelength , final_intensity = get_non_repeated_wavelength(detected_wavelength, detected_wavelength_intensity)


start = time.time()
sorted_final_wavelength = sorted(final_wavelength, reverse=True)
sorted_final_intensity = []
for i in range(len(sorted_final_wavelength)):
    for j in range(len(final_wavelength)):
        if (sorted_final_wavelength[i] == final_wavelength[j]):
            sorted_final_intensity.append(final_intensity[j])


intensity = []
for i in range(len(wavelength)):
    for j in range(len(sorted_final_wavelength)):
        if (wavelength[i]==sorted_final_wavelength[j]):
            intensity.append(sorted_final_intensity[j])
            result = 1
            break
        else:
            result = 0

    if (result==0):
        intensity.append(0)
        result=1
stop = time.time()
print("time required to sort wavelength and intensity = ", stop - start)


plt.plot(wavelength, intensity)
plt.xlabel("(VIBGYOR) wavelength (nm) --->")
plt.ylabel("magnitude/intensity")
plt.show()

total_stop = time.time()
print("total time elapsed = ", total_stop - total_start)

#storing detected wavelength to xls
xl_path = "pixel_wavelength.xlsx"
wb_obj = openpyxl.load_workbook(xl_path)
sheet_obj = wb_obj.active
l= len(detected_wavelength)
for i in range(l):
    cell_obj4 = sheet_obj.cell(row=1 + i, column=6)
    cell_obj4.value = str(detected_wavelength[i])
    cell_obj5 = sheet_obj.cell(row=1 + i, column=7)
    cell_obj5.value = str(detected_wavelength_intensity[i])
    wb_obj.save(xl_path)


