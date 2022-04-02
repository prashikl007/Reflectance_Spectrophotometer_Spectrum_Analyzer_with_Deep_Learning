import cv2
import time
import numpy as np
import matplotlib.pyplot as plt
import openpyxl



#reading image using cv2
img_path = 'C:/Users/Public/PYTHON_PROJECTS/spectrometer/spectrum_2.jpg'
image = cv2.imread(img_path)
cv2.imshow("image", image)
cv2.waitKey(0)
cv2.destroyAllWindows()

h, w, _ = image.shape
print("height = ", h , ", width = ", w)

#ordering BGR to RGB from original image into img_rgb
img_rgb = []
start = time.time()
for i in range(h):
    img_rgb.append([])
    for j in range(w):
        img_rgb[i].append([image[i][j][2], image[i][j][1], image[i][j][0]])
stop = time.time()
print("time required for ordering = ", stop-start)



#plotting RGB
r = []
g = []
b = []
for i in range(w):
    x = 0
    y = 0
    z = 0
    for j in range(h):
        x = x + img_rgb[j][i][0]
        y = y + img_rgb[j][i][1]
        z = z + img_rgb[j][i][2]
    r.append(x/h)
    g.append(y/h)
    b.append(z/h)


plt.plot(r, color='red')
plt.plot(g, color='green')
plt.plot(b, color='blue')
plt.show()


#getting colors value
color_pixel = []
for i in range(w):
    color_pixel.append(img_rgb[0][i])

#calculating steps by which nm wavelength should increase
step_size = (700-w)/w

#generating wavelength for each color pixel
wavelength = []
for i in range(w):
    wavelength.append(700-i*step_size)


xl_path = "pixel_wavelength.xlsx"
wb_obj = openpyxl.load_workbook(xl_path)

sheet_obj = wb_obj.active

for i in range(w):
    cell_obj1 = sheet_obj.cell(row = 1+i, column = 1)
    cell_obj2 = sheet_obj.cell(row=1+i, column=2)
    cell_obj3 = sheet_obj.cell(row=1 + i, column=3)
    cell_obj4 = sheet_obj.cell(row=1 + i, column=4)
    cell_obj1.value = str(color_pixel[i][0])
    cell_obj2.value = str(color_pixel[i][1])
    cell_obj3.value = str(color_pixel[i][2])
    cell_obj4.value = str(wavelength[i])
    wb_obj.save(xl_path)




row = sheet_obj.max_row
column = sheet_obj.max_column

step_size = (700-row)/row
for i in range(row):
    cell_obj4 = sheet_obj.cell(row=1 + i, column=4)
    cell_obj4.value = str(700-i*step_size)
    wb_obj.save(xl_path)


